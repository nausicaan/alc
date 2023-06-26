#!/usr/bin/env python3

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink

REPLACEMENTS = [
    ("http:", "https:"),
    ("Docs\n", ""),
    ("click here\n", ""),
    ("Click here\n", ""),
    ("Decision\n", ""),
    ("None\n", ""),
    ("N/A\n", "")
]

home_dir = str(Path.home())
regions = ("interior-", "island-", "kootenays-", "north-", "okanagan-", "south_coast-")
years = ("2006", "2007", "2008", "2009", "2010", "2011", "2012", "2013", "2014", "2015", "2016")
fpath = home_dir + "/Documents/alc/files/spreadsheets/"
folder = "-decision-minutes/"
urls = ""


# Gather the urls from hidden columns used in the 2006 -> 2014 editions
def iterating_column(path, sheet_name, col):
    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return

    sheet = workbook[sheet_name]
    for cell in sheet[col]:
        global urls
        urls = urls + f"{cell.value}" + "\n"


# Gather the urls from embeded links used in the 2015 and 2016 editions
def edge_cases(route):
    workbook = load_workbook(route)
    sheet = workbook["Sheet1"]
    for u in sheet['F'][1:]:
        if u.value is None:
            continue
            [(_, display)] = [p for p in u.hyperlink if 'display' in p]
            [(_, ref)] = [p for p in u.hyperlink if 'ref' in p]
            location = None
            tooltip = None
            wid: object
            [(_, wid)] = [p for p in u.hyperlink if 'id' in p]
            target = None

            if "DownloadAsset" in display:
                u.hyperlink = Hyperlink(
                    display=display,
                    ref=ref,
                    location=location,
                    tooltip=tooltip,
                    id=wid,
                    target=target)
        global urls
        if u.hyperlink is None:
            continue
        elif len(str(u.hyperlink.display)) < 100:
            continue
        else:
            urls = urls + str(u.hyperlink.display) + "\n"


# Iterate through the spreadsheets chronologically and by region, compensating for the edge cases
if __name__ == "__main__":
    for y in years:
        if int(y) < 2015:
            for r in regions:
                if r == "south_coast-":
                    iterating_column(fpath + y + folder + r + y + ".xlsx", "South Coast-" + y, "E")
                else:
                    iterating_column(fpath + y + folder + r + y + ".xlsx", r.title() + y, "E")
        else:
            for r in regions:
                if r == "south_coast-" and y == "2016":
                    edge_cases(fpath + y + folder + "south-coast-2016.xlsx")
                else:
                    edge_cases(fpath + y + folder + r + y + ".xlsx")

    for old, new in REPLACEMENTS:
        urls = urls.replace(old, new)

    f = open("sources/possible.txt", "w")
    f.write(urls)
    f.close()
